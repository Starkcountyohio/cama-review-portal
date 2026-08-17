"""
update_prompt.py
────────────────────────────────────────────────────────────────
Stark County Auditor — Monthly AI Prompt Retraining Script

Run this once a month after collecting new appraiser decisions:
    python prompt_training/update_prompt.py

Requirements:
    - data/Grade_Con_Compara.xlsx     (appraiser final Grade/CDU decisions)
    - data/weekly_json/               (folder of 2026-WNN.json files)
    - review_portal_template.html     (in parent folder)

What it does:
    1. Loads all weekly JSON parcel data (MLS remarks, YRBLT, CDU, etc.)
    2. Matches against appraiser final decisions from Grade_Con_Compara.xlsx
    3. Runs statistical analysis on what remark patterns predicted changes
    4. Rewrites the AI prompt section in review_portal_template.html
    5. Prints a summary report of what changed
"""

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── Paths ──────────────────────────────────────────────────────────────────────
ROOT         = Path(__file__).parent.parent
TEMPLATE     = ROOT / "review_portal_template.html"
COMPARA_FILE = ROOT / "data" / "Grade_Con_Compara.xlsx"
JSON_DIR     = ROOT / "data" / "weekly_json"

# ── Grade / CDU ranking for direction detection ────────────────────────────────
GRADE_RANK = {'E':0,'D':1,'C-':2,'C':3,'C+':4,'B-':5,'B':6,'B+':7,
              'A-':8,'A':9,'A+':10,'X-':11,'X':12,'X+':13}
CDU_RANK   = {'PR':0,'FR':1,'AV':2,'GD':3,'VG':4,'EX':5}

# ── Keyword groups to analyze ──────────────────────────────────────────────────
UPGRADE_PHRASES = [
    'new roof','newer roof','new furnace','new hvac','new a/c','new windows',
    'updated kitchen','updated bath','renovated','remodeled','new floors',
    'new electric','new plumbing','new water heater','new siding',
    'move-in ready','move in ready','updated throughout','completely updated',
    'fully updated','immaculate','meticulously','turnkey',
]
DOWNGRADE_PHRASES = [
    'tlc','opportunity','as-is','as is','investor','handyman','needs work',
    'needs updating','fixer','sweat equity','elbow grease','potential',
    'estate sale','cash only','sold as is','sold as-is','vacant','abandoned',
]
NOISE_PHRASES = [
    'great location','convenient location','minutes from','close to',
    'near shopping','desirable','sought after','award winning school',
    'location location',
]

# ── Loaders ────────────────────────────────────────────────────────────────────

def load_compara(path: Path) -> dict:
    """Load Grade_Con_Compara.xlsx → {parid: {grade, cdu}}"""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    result = {}
    for row in rows[1:]:
        if not row[0]:
            continue
        parid = str(row[0]).strip()
        result[parid] = {'grade': str(row[1]).strip(), 'cdu': str(row[2]).strip()}
    return result


def load_weekly_parcels(json_dir: Path) -> dict:
    """Load all weekly JSON files → {parid: parcel_dict}"""
    parcels = {}
    json_files = sorted(json_dir.glob("*.json"))
    if not json_files:
        print(f"WARNING: No JSON files found in {json_dir}")
        return parcels

    for jf in json_files:
        with open(jf) as f:
            data = json.load(f)
        week = jf.stem
        rows = data.get('mismatches', []) + data.get('perfects', [])
        for p in rows:
            pid = str(p.get('Parcel_ID', '')).strip()
            if pid and pid not in parcels:
                parcels[pid] = dict(p, _week=week)
    return parcels


# ── Analysis ───────────────────────────────────────────────────────────────────

def build_matched_dataset(compara: dict, portal: dict) -> list:
    """Join compara decisions with portal parcel data."""
    matched = []
    unmatched = 0
    for parid, final in compara.items():
        p = portal.get(parid)
        if not p:
            unmatched += 1
            continue
        orig_grade = str(p.get('GRADE', '')).strip()
        orig_cdu   = str(p.get('CDU', '')).strip()
        fin_grade  = final['grade']
        fin_cdu    = final['cdu']
        matched.append({
            'parcelId':     parid,
            'orig_grade':   orig_grade,
            'orig_cdu':     orig_cdu,
            'final_grade':  fin_grade,
            'final_cdu':    fin_cdu,
            'grade_changed': orig_grade != fin_grade,
            'cdu_changed':   orig_cdu   != fin_cdu,
            'remarks':      str(p.get('Public_Remarks', '')).lower(),
            'yrblt':        p.get('YRBLT'),
            '_week':        p.get('_week', ''),
        })
    if unmatched:
        print(f"  WARNING: {unmatched} parcels in compara not found in weekly JSON data")
    return matched


def analyze(matched: list) -> dict:
    """Run full statistical analysis. Returns findings dict."""
    total       = len(matched)
    grade_chg   = [m for m in matched if m['grade_changed']]
    cdu_chg     = [m for m in matched if m['cdu_changed']]
    either_chg  = [m for m in matched if m['grade_changed'] or m['cdu_changed']]

    # CDU change directions
    upgrades   = [m for m in cdu_chg if CDU_RANK.get(m['final_cdu'],2) > CDU_RANK.get(m['orig_cdu'],2)]
    downgrades = [m for m in cdu_chg if CDU_RANK.get(m['final_cdu'],2) < CDU_RANK.get(m['orig_cdu'],2)]

    # Common CDU transitions
    cdu_transitions = Counter(f"{m['orig_cdu']}→{m['final_cdu']}" for m in cdu_chg).most_common(8)
    grade_transitions = Counter(f"{m['orig_grade']}→{m['final_grade']}" for m in grade_chg).most_common(8)

    # Upgrade phrase rates
    upgrade_stats = []
    for phrase in UPGRADE_PHRASES:
        in_upgrades = sum(1 for m in upgrades if phrase in m['remarks'])
        in_total    = sum(1 for m in matched  if phrase in m['remarks'])
        if in_total < 3:
            continue
        baseline = in_total / total
        rate     = in_upgrades / len(upgrades) if upgrades else 0
        multiplier = rate / baseline if baseline > 0 else 0
        upgrade_stats.append({
            'phrase': phrase, 'count': in_upgrades, 'total': in_total,
            'rate': rate, 'baseline': baseline, 'multiplier': multiplier
        })
    upgrade_stats.sort(key=lambda x: -x['multiplier'])

    # Downgrade phrase rates
    downgrade_stats = []
    for phrase in DOWNGRADE_PHRASES:
        in_downgrades = sum(1 for m in downgrades if phrase in m['remarks'])
        in_total      = sum(1 for m in matched   if phrase in m['remarks'])
        if in_total < 2:
            continue
        baseline   = in_total / total
        rate       = in_downgrades / len(downgrades) if downgrades else 0
        multiplier = rate / baseline if baseline > 0 else 0
        downgrade_stats.append({
            'phrase': phrase, 'count': in_downgrades, 'total': in_total,
            'rate': rate, 'baseline': baseline, 'multiplier': multiplier
        })
    downgrade_stats.sort(key=lambda x: -x['multiplier'])

    # Hard negative change rate
    hard_neg_kws = ['estate sale','cash only','sold as is','sold as-is','vacant','abandoned']
    hard_neg = [m for m in matched if any(kw in m['remarks'] for kw in hard_neg_kws)]
    hard_neg_change_rate = sum(1 for m in hard_neg if m['cdu_changed']) / len(hard_neg) if hard_neg else 0

    # Noise confirmation
    noise_in_changed   = sum(1 for m in either_chg if any(kw in m['remarks'] for kw in NOISE_PHRASES))
    noise_in_unchanged = sum(1 for m in matched if not (m['grade_changed'] or m['cdu_changed'])
                             and any(kw in m['remarks'] for kw in NOISE_PHRASES))
    noise_changed_pct   = noise_in_changed / len(either_chg) * 100 if either_chg else 0
    noise_unchanged_pct = noise_in_unchanged / (total - len(either_chg)) * 100 if total - len(either_chg) else 0

    # YRBLT medians
    import statistics
    yrblt_upgraded = [m['yrblt'] for m in upgrades
                      if isinstance(m.get('yrblt'), (int, float)) and m['yrblt'] and m['yrblt'] > 1800]
    yrblt_median_upgraded = int(statistics.median(yrblt_upgraded)) if yrblt_upgraded else 0

    return {
        'total': total,
        'grade_change_pct': len(grade_chg) / total * 100,
        'cdu_change_pct':   len(cdu_chg)   / total * 100,
        'either_change_pct': len(either_chg) / total * 100,
        'n_upgrades':   len(upgrades),
        'n_downgrades': len(downgrades),
        'cdu_transitions':   cdu_transitions,
        'grade_transitions': grade_transitions,
        'upgrade_stats':     upgrade_stats[:12],
        'downgrade_stats':   downgrade_stats[:10],
        'hard_neg_n':          len(hard_neg),
        'hard_neg_change_pct': hard_neg_change_rate * 100,
        'noise_changed_pct':   noise_changed_pct,
        'noise_unchanged_pct': noise_unchanged_pct,
        'yrblt_median_upgraded': yrblt_median_upgraded,
    }


# ── Prompt builder ─────────────────────────────────────────────────────────────

def build_prompt(f: dict) -> str:
    """Build the full AI prompt string from analysis findings."""

    # Top CDU transitions as readable string
    transitions_str = ', '.join(f"{t} ({n}x)" for t, n in f['cdu_transitions'][:6])

    # Top upgrade signals
    upgrade_lines = []
    for s in f['upgrade_stats']:
        if s['multiplier'] >= 1.5 and s['count'] >= 3:
            upgrade_lines.append(
                f'- "{s["phrase"]}" → {s["count"]}/{f["n_upgrades"]} upgrades '
                f'({s["rate"]*100:.0f}%, {s["multiplier"]:.1f}x baseline)'
            )

    # Top downgrade signals
    downgrade_lines = []
    for s in f['downgrade_stats']:
        if s['multiplier'] >= 1.5 and s['count'] >= 2:
            downgrade_lines.append(
                f'- "{s["phrase"]}" → {s["count"]}/{f["n_downgrades"]} downgrades '
                f'({s["rate"]*100:.0f}%, {s["multiplier"]:.1f}x baseline)'
            )

    upgrade_block   = '\n'.join(upgrade_lines)   or '- (insufficient data)'
    downgrade_block = '\n'.join(downgrade_lines) or '- (insufficient data)'

    return f"""const textPrompt = `You are a certified property appraiser assistant for the Stark County Auditor's Office in Ohio. Suggest a Grade and Condition based on all available evidence.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
GRADE SCALE (2024 Stark County)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
X+/X/X-  Excellent   — architect-designed, exceptional custom materials
A+/A/A-  Very Good   — high-quality, custom ornamentation, premium finishes
B+/B/B-  Good        — above-average, upgraded materials
C+/C/C-  Average     — mass-produced, meets code, stock materials
D        Fair        — below-average materials, plain interior
E        Low         — basic minimum code

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CONDITION SCALE (1–5)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1 Excellent  — near new, 10–12 pluses
2 Good       — well maintained, 4–9 pluses
3 Average    — normal wear, 0–3 pluses/minuses
4 Fair       — deferred maintenance, 4–9 minuses
5 Poor       — major repairs needed, 10–12 minuses

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
STARK COUNTY EMPIRICAL DATA — {f['total']} REVIEWED PARCELS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
These are ACTUAL appraiser decisions from this office. Weight heavily.

CHANGE RATES:
- Appraisers changed CDU on {f['cdu_change_pct']:.0f}% of parcels, Grade on {f['grade_change_pct']:.0f}%
- {f['either_change_pct']:.0f}% had at least one change; {100-f['either_change_pct']:.0f}% confirmed unchanged
- Most common CDU changes: {transitions_str}
- When uncertain between two CDU values lean toward GD or AV — these are the most common outcomes

HARD NEGATIVE LANGUAGE = STRONGEST SIGNAL:
- "Estate sale / cash only / sold as-is / vacant / abandoned" → {f['hard_neg_change_pct']:.0f}% CDU change rate (n={f['hard_neg_n']})
- These phrases = almost certainly FR or PR. Do not soften this.

CDU UPGRADE SIGNALS (ranked by Stark County predictive power):
{upgrade_block}

CDU DOWNGRADE SIGNALS (ranked by Stark County predictive power):
{downgrade_block}

YEAR BUILT: Homes with median YRBLT {f['yrblt_median_upgraded']} CAN be upgraded if truly renovated — age alone does not determine CDU.

LOCATION/SCHOOL FLUFF — IGNORE COMPLETELY:
"Great location / award-winning schools / minutes from shopping / desirable / sought-after"
appeared at {f['noise_changed_pct']:.0f}% in changed vs {f['noise_unchanged_pct']:.0f}% in unchanged parcels — zero predictive value.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
EVIDENCE WEIGHTING HIERARCHY
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. PHOTO — highest weight. Roof, siding, windows, overall upkeep. If photo shows deferred maintenance, trust it over ANY positive remarks.
2. HARD NEGATIVE LANGUAGE — nearly always FR/PR regardless of other language
3. SPECIFIC NAMED SYSTEM UPDATES — remodeled/renovated/new furnace/new windows/updated kitchen+bath in combination
4. SOFT NEGATIVE LANGUAGE — opportunity/investor/as-is = strong lean toward FR
5. POSITIVE GENERIC LANGUAGE — lowest weight. "Beautifully updated/move-in ready" appeared on AV and FR properties. Do NOT use alone.
6. COSMETIC ONLY — fresh paint/new carpet/stainless/hardwood alone = no CDU change warranted

PHOTO vs REMARKS CONFLICT: If photo shows worn exterior but remarks claim "move-in ready" — TRUST THE PHOTO. Flag the conflict in rationale.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PROPERTY DATA
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Address: \${{parcel.address}}, \${{parcel.city}} OH \${{parcel.zip}}
Year Built: \${{parcel.yrblt || "unknown"}}
Above-Grade Sq Ft (SFLA): \${{parcel.sfla != null ? parcel.sfla : "unknown"}}
\${{bathStr ? \`Bathrooms: \${{bathStr}}\` : ""}}
Current CAMA Grade: \${{parcel.grade || "unknown"}}
Current CAMA CDU: \${{parcel.cdu || "unknown"}}
MLS Public Remarks: \${{parcel.remarks || "(none provided)"}}
\${{hasPhoto ? "Front exterior photo attached — weight this heavily." : "No photo — reduce confidence, rely on remarks and year built only."}}

Respond ONLY with valid JSON — no markdown, no extra text:
{{
  "suggestedGrade": "e.g. B+",
  "suggestedCondition": 3,
  "conditionLabel": "e.g. Average",
  "gradeConfidence": "Low|Moderate|High",
  "conditionConfidence": "Low|Moderate|High",
  "rationale": "2–3 sentences. Cite specific evidence used, flag realtor bias detected, note any photo vs. remarks conflict."
}}\`;"""


# ── Template updater ───────────────────────────────────────────────────────────

def update_template(template_path: Path, new_prompt: str) -> bool:
    """Replace the AI prompt in review_portal_template.html."""
    with open(template_path, 'r', encoding='utf-8') as f:
        html = f.read()

    start = html.find('const textPrompt')
    if start == -1:
        print("ERROR: Could not find 'const textPrompt' in template")
        return False

    end = html.find('`;', start) + 2
    if end < 2:
        print("ERROR: Could not find end of prompt (`;)")
        return False

    html = html[:start] + new_prompt + html[end:]

    with open(template_path, 'w', encoding='utf-8') as f:
        f.write(html)
    return True


# ── Report printer ─────────────────────────────────────────────────────────────

def print_report(f: dict, weeks: list):
    print("\n" + "="*65)
    print("  STARK COUNTY PROMPT RETRAINING REPORT")
    print("="*65)
    print(f"  Weeks analyzed:    {', '.join(weeks)}")
    print(f"  Total parcels:     {f['total']}")
    print(f"  Grade change rate: {f['grade_change_pct']:.0f}%")
    print(f"  CDU change rate:   {f['cdu_change_pct']:.0f}%")
    print(f"\n  Top CDU transitions:")
    for t, n in f['cdu_transitions'][:5]:
        print(f"    {t}: {n}x")
    print(f"\n  Top upgrade signals:")
    for s in f['upgrade_stats'][:5]:
        if s['multiplier'] >= 1.5:
            print(f"    \"{s['phrase']}\": {s['multiplier']:.1f}x baseline")
    print(f"\n  Top downgrade signals:")
    for s in f['downgrade_stats'][:5]:
        if s['multiplier'] >= 1.5:
            print(f"    \"{s['phrase']}\": {s['multiplier']:.1f}x baseline")
    print(f"\n  Hard negative change rate: {f['hard_neg_change_pct']:.0f}% (n={f['hard_neg_n']})")
    print("="*65)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("\nStark County — Monthly Prompt Retraining")
    print("-"*45)

    # Validate files exist
    for path, label in [(COMPARA_FILE, "Grade_Con_Compara.xlsx"),
                        (TEMPLATE,     "review_portal_template.html"),
                        (JSON_DIR,     "data/weekly_json/")]:
        if not path.exists():
            print(f"ERROR: Missing {label} at {path}")
            print("       Check your folder structure matches CLAUDE.md")
            sys.exit(1)

    # Load data
    print("\nLoading appraiser decisions...")
    compara = load_compara(COMPARA_FILE)
    print(f"  {len(compara)} parcel decisions loaded")

    print("Loading weekly parcel data...")
    portal  = load_weekly_parcels(JSON_DIR)
    print(f"  {len(portal)} unique parcels loaded")

    weeks = sorted(set(p['_week'] for p in portal.values() if p.get('_week')))
    print(f"  Weeks: {', '.join(weeks)}")

    # Match and analyze
    print("\nMatching and analyzing...")
    matched = build_matched_dataset(compara, portal)
    print(f"  Matched {len(matched)}/{len(compara)} parcels")

    if len(matched) < 50:
        print(f"WARNING: Only {len(matched)} matched parcels — results may not be reliable")
        ans = input("Continue anyway? (y/n): ").strip().lower()
        if ans != 'y':
            sys.exit(0)

    findings = analyze(matched)
    print_report(findings, weeks)

    # Build and inject new prompt
    print("\nUpdating AI prompt in template...")
    new_prompt = build_prompt(findings)
    success = update_template(TEMPLATE, new_prompt)

    if success:
        print(f"  ✓ Prompt updated in {TEMPLATE.name}")
        print(f"\nNext steps:")
        print(f"  1. Run build_portal.py to rebuild review_portal.html with new prompt")
        print(f"  2. Upload review_portal.html to GitHub")
        print(f"  3. Staff will get improved AI suggestions immediately\n")
    else:
        print("  ✗ Failed to update template — check errors above")
        sys.exit(1)

    input("Press Enter to close...")


if __name__ == "__main__":
    main()
